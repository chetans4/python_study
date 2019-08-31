#Demo Project

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    worksheet = wb['Sheet1']
    #cell = worksheet['a1']
    #cell = worksheet.cell(1,2)
    #print cell.value
    #print worksheet.max_row


    for row in range(2, worksheet.max_row +1):
        cell = worksheet.cell(row, 3);
        #print cell.value
        corrected_value = cell.value *0.9;
        corrected_value_cell = worksheet.cell(row, 4)
        corrected_value_cell.value = corrected_value;
        

    values = Reference(worksheet, min_row=2, max_row = worksheet.max_row, min_col= 4, max_col =4)

    chart = BarChart();
    chart.add_data(values)

    worksheet.add_chart(chart, "a8");

    wb.save(filename);

    print "Completed!"
    
process_workbook("Book1_corrected.xlsx")